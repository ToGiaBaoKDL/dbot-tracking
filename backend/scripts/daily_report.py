#!/usr/bin/env python3
"""Generate daily signals report (xlsx) and send via email.

Usage:
    python scripts/daily_report.py [--date YYYY-MM-DD]

Env:
    DATABASE_URL
    SMTP_HOST (default: smtp.gmail.com)
    SMTP_PORT (default: 465)
    SMTP_USER
    SMTP_PASSWORD
    EMAIL_FROM
    EMAIL_TO (comma-separated)
"""

import argparse
import logging
import os
import smtplib
import sys
import tempfile
from datetime import date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

sys.path.insert(0, "/app")

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.core.db_url import to_sync_url

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("dbot-etl")

# ── Config ──
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "")
EMAIL_TO = os.environ.get("EMAIL_TO", "")

# ── Excel styling ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BUY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BUY_FONT = Font(color="006100")
SELL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SELL_FONT = Font(color="9C0006")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_db_session():
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL is required")
    sync_url = to_sync_url(db_url)
    engine = create_engine(sync_url, pool_pre_ping=True)
    session_cls = sessionmaker(bind=engine)
    return session_cls()


def fetch_signals(session, target_date: date):
    """Return (buy_rows, sell_rows) as list of dicts."""
    sql = text("""
        SELECT symbol, record_date, close_price, volume, signal
        FROM stock_daily_data
        WHERE record_date = :d AND signal IN ('BUY', 'SELL')
        ORDER BY signal, symbol
    """)
    rows = session.execute(sql, {"d": target_date}).mappings().all()
    buy = [r for r in rows if r["signal"] == "BUY"]
    sell = [r for r in rows if r["signal"] == "SELL"]
    return buy, sell


def create_xlsx(target_date: date, buy_rows, sell_rows, filepath: str):
    wb = openpyxl.Workbook()
    headers = ["Mã CK", "Ngày", "Giá đóng cửa", "Khối lượng", "Tín hiệu"]

    def _write_sheet(ws, title, rows, fill, font_color):
        ws.title = title
        # Header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        # Data
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row["symbol"])
            ws.cell(row=row_idx, column=2, value=str(row["record_date"]))
            ws.cell(row=row_idx, column=3, value=float(row["close_price"]) if row["close_price"] else None)
            ws.cell(row=row_idx, column=4, value=int(row["volume"]) if row["volume"] else None)
            ws.cell(row=row_idx, column=5, value=row["signal"])
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = fill
                cell.font = font_color
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center" if col != 1 else "left", vertical="center")
        # Auto-width
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.freeze_panes = "A2"

    _write_sheet(wb.active, "MUA", buy_rows, BUY_FILL, BUY_FONT)
    ws_sell = wb.create_sheet("BÁN")
    _write_sheet(ws_sell, "BÁN", sell_rows, SELL_FILL, SELL_FONT)

    wb.save(filepath)
    logger.info("Report saved: %s", filepath)


def send_email(target_date: date, buy_rows, sell_rows, attachment_path: str):
    if not all([SMTP_USER, SMTP_PASSWORD, EMAIL_FROM, EMAIL_TO]):
        logger.warning("Email config incomplete, skipping notification")
        return

    recipients = [e.strip() for e in EMAIL_TO.split(",") if e.strip()]
    if not recipients:
        logger.warning("No recipients configured")
        return

    subject = f"Báo cáo tín hiệu DBOT — {target_date.strftime('%d/%m/%Y')}"
    body = (
        f"Xin chào,\n\n"
        f"Báo cáo tín hiệu mua/bán ngày {target_date.strftime('%d/%m/%Y')} đã được đính kèm.\n\n"
        f"- Sheet MUA: {len(buy_rows)} mã\n"
        f"- Sheet BÁN: {len(sell_rows)} mã\n\n"
        f"Trân trọng,\nDBOT Signals Tracker"
    )

    msg = MIMEMultipart()
    msg["From"] = EMAIL_FROM
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    # Attachment
    filename = Path(attachment_path).name
    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        logger.info("Email sent to %s", ", ".join(recipients))
    except Exception as e:
        logger.error("Failed to send email: %s", e)
        raise


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", type=date.fromisoformat, default=date.today())
    args = parser.parse_args()
    target_date = args.date

    logger.info("Generating daily report for %s", target_date)

    session = get_db_session()
    try:
        buy_rows, sell_rows = fetch_signals(session, target_date)
        logger.info("Signals: %s BUY, %s SELL", len(buy_rows), len(sell_rows))
    finally:
        session.close()

    if not buy_rows and not sell_rows:
        logger.info("No signals for %s, skipping report", target_date)
        return

    tmp_dir = tempfile.mkdtemp(prefix="dbot_report_")
    filepath = os.path.join(tmp_dir, f"dbot_signals_{target_date.isoformat()}.xlsx")

    try:
        create_xlsx(target_date, buy_rows, sell_rows, filepath)
        send_email(target_date, buy_rows, sell_rows, filepath)
    finally:
        # Clean up temp file
        try:
            os.remove(filepath)
            os.rmdir(tmp_dir)
        except OSError:
            pass


if __name__ == "__main__":
    main()
